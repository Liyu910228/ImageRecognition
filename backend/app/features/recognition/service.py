from __future__ import annotations

import shutil
import time
import uuid
from pathlib import Path
from tempfile import NamedTemporaryFile
from datetime import datetime
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.core.config import settings
from app.features.recognition.image_preprocess import crop_bottom_watermark
from app.features.recognition.logs import append_recognition_log
from app.features.recognition.matcher import SUPPORTED_EXTENSIONS, match_uploaded_image
from app.features.recognition.product_tags import (
    competitor_brand_matches,
    has_competitor_brand,
    is_weak_edge_product_evidence,
    sanitize_recognition_text,
    split_brand_config,
    snow_brand_matches,
)
from app.features.settings.service import get_model_settings
from app.services.qwen_client import analyze_product_image


RETRIABLE_IMAGE_DOWNLOAD_STATUS = {429, 500, 502, 503, 504}
VISION_FALLBACK_SOURCES = {"no_api_key", "model_error", "model_timeout", "qwen_error", "qwen_timeout"}


def _validate_upload(file: UploadFile) -> str:
    filename = file.filename or ""
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 JPG、PNG、WebP 图片")
    return suffix


def _validate_workbook(file: UploadFile) -> str:
    filename = file.filename or ""
    suffix = Path(filename).suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 表格")
    return suffix


def _ensure_index_ready() -> None:
    if not settings.vector_index_path.exists():
        raise HTTPException(status_code=409, detail="模板向量索引不存在，请先重建索引")


def _split_hints(hints: str) -> list[str]:
    return [item.strip() for item in hints.replace("，", " ").replace(",", " ").split() if item.strip()]


def _normalize_header(value: object) -> str:
    return str(value or "").replace(" ", "").replace("\n", "").strip().lower()


def _is_photo_url_header(value: object) -> bool:
    header = _normalize_header(value)
    return header == "照片链接" or ("链接" in header and ("照片" in header or "图片" in header or "photo" in header or "image" in header))


def _is_http_url(value: object) -> bool:
    text = str(value or "").strip()
    return text.startswith("http://") or text.startswith("https://")


def extract_image_links_from_workbook(file: UploadFile) -> dict[str, object]:
    _validate_workbook(file)
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        shutil.copyfileobj(file.file, temp_file)
        temp_path = Path(temp_file.name)

    workbook = None
    try:
        workbook = load_workbook(temp_path, read_only=True, data_only=True)
        links: list[dict[str, object]] = []
        matched_header = ""
        for sheet in workbook.worksheets:
            header_row_index = 0
            photo_column_index = -1
            for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True), start=1):
                for column_index, value in enumerate(row):
                    if _is_photo_url_header(value):
                        header_row_index = row_index
                        photo_column_index = column_index
                        matched_header = str(value or "").strip()
                        break
                if photo_column_index >= 0:
                    break
            if photo_column_index < 0:
                continue

            for row_index, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
                if photo_column_index >= len(row):
                    continue
                image_url = str(row[photo_column_index] or "").strip()
                if _is_http_url(image_url):
                    links.append({"row": row_index, "sheet": sheet.title, "image_url": image_url})

        if not links:
            raise HTTPException(status_code=400, detail="表格中没有找到“照片链接”字段或有效图片链接")
        return {
            "filename": file.filename or temp_path.name,
            "field": matched_header or "照片链接",
            "count": len(links),
            "links": links,
        }
    finally:
        if workbook is not None:
            workbook.close()
        temp_path.unlink(missing_ok=True)


def _recognize_saved_image(
    upload_path: Path,
    filename: str,
    hints: str = "",
    api_key_override: str | None = None,
    trace_id: str = "",
    model_profile: str = "default",
) -> dict[str, object]:
    _ensure_index_ready()
    effective_trace_id = _normalize_trace_id(trace_id) or _generate_trace_id()
    recognition_path = crop_bottom_watermark(upload_path)
    footer_cropped = recognition_path != upload_path
    analysis, vision_retried = _analyze_product_image_with_retry(
        recognition_path,
        api_key_override=api_key_override,
        model_profile=model_profile,
    )
    model_settings = get_model_settings()
    if analysis.source in VISION_FALLBACK_SOURCES:
        return _recognize_with_local_visual_fallback(
            recognition_path=recognition_path,
            filename=filename,
            trace_id=effective_trace_id,
            hints=hints,
            analysis_source=analysis.source,
            api_key_source=analysis.api_key_source,
            footer_cropped=footer_cropped,
            vision_retried=vision_retried,
            model_calls=analysis.model_calls or [],
            model_settings=model_settings,
        )
    cleaned_analysis_texts = [
        sanitize_recognition_text(value)
        for value in [
            *analysis.brands,
            *analysis.keywords,
            *analysis.product_text,
            analysis.primary_product_description,
        ]
    ]
    cleaned_analysis_texts = [value for value in cleaned_analysis_texts if value]
    keywords = [
        *cleaned_analysis_texts,
        *_split_hints(hints),
    ]
    matched_snow_brands = snow_brand_matches(
        *cleaned_analysis_texts,
        snow_brands=split_brand_config(model_settings.get("snow_brand_names", "")),
    )
    image_has_snow_brand = bool(matched_snow_brands)
    matched_competitor_brands = competitor_brand_matches(*cleaned_analysis_texts)
    image_has_competitor_brand = bool(matched_competitor_brands) or has_competitor_brand(*cleaned_analysis_texts)
    weak_edge_product = is_weak_edge_product_evidence(*cleaned_analysis_texts)
    if (not image_has_snow_brand) or weak_edge_product:
        gate_reason = "只识别到边缘/模糊小商品，直接未命中" if weak_edge_product else "未提取到雪花体系品牌，直接未命中"
        result = {
            "filename": filename,
            "traceId": effective_trace_id,
            "best": None,
            "candidates": [],
            "reviewRequired": False,
            "analysis": {
                "source": analysis.source,
                "packageType": analysis.package_type,
                "keywords": analysis.keywords,
                "brands": analysis.brands,
                "productText": analysis.product_text,
                "isMultiProduct": analysis.is_multi_product,
                "primaryProductDescription": analysis.primary_product_description,
                "manualHints": _split_hints(hints),
                "apiKeySource": analysis.api_key_source,
                "modelProfile": analysis.model_profile,
                "modelCalls": analysis.model_calls or [],
                "footerCropped": footer_cropped,
                "visionRetried": vision_retried,
                "matchedSnowBrands": matched_snow_brands,
                "matchedCompetitorBrands": matched_competitor_brands,
            },
            "pipeline": {
                "stage1": {
                    "name": "雪花品牌门禁",
                    "packageType": analysis.package_type,
                    "candidateCount": 0,
                    "fallback": False,
                },
                "stage2": {
                    "name": "品牌和关键词",
                    "keywords": [gate_reason],
                    "candidateCount": 0,
                    "fallback": False,
                },
                "stage3": {
                    "name": "图片语义相似度",
                    "candidateCount": 0,
                },
            },
            "models": {
                "embeddingModel": model_settings["embedding_model"],
                "vlModel": model_settings["vl_model"],
                "enableVlRerank": model_settings["enable_vl_rerank"],
                "qwenTimeoutSeconds": model_settings["qwen_timeout_seconds"],
            },
        }
        _write_recognition_log(
            filename=filename,
            trace_id=effective_trace_id,
            result=result,
            analysis_source=analysis.source,
            cleaned_texts=cleaned_analysis_texts,
            gate_reason=gate_reason,
            image_has_snow_brand=image_has_snow_brand,
            image_has_competitor_brand=image_has_competitor_brand,
            matched_snow_brands=matched_snow_brands,
            matched_competitor_brands=matched_competitor_brands,
            model_calls=analysis.model_calls or [],
            footer_cropped=footer_cropped,
        )
        return result
    ranking_weights = {
        key: float(model_settings[key])
        for key in [
            "weight_brand_match_visual",
            "weight_brand_match_brand",
            "weight_brand_match_text",
            "weight_brand_miss_visual",
            "weight_brand_miss_text",
            "weight_no_brand_visual",
            "weight_no_brand_text",
        ]
    }
    matches, pipeline = match_uploaded_image(
        upload_path=recognition_path,
        index_path=settings.vector_index_path,
        template_images_dir=settings.template_images_dir,
        top_k=int(model_settings.get("top_k", 5)),
        package_type=analysis.package_type,
        brands=matched_snow_brands,
        keywords=[keyword for keyword in keywords if keyword],
        weights=ranking_weights,
        strict_brand_filter=True,
    )
    threshold = float(model_settings["low_confidence_threshold"])
    best = matches[0] if matches else None
    result = {
        "filename": filename,
        "traceId": effective_trace_id,
        "best": best.__dict__ if best else None,
        "candidates": [match.__dict__ for match in matches],
        "reviewRequired": bool((best and best.score < threshold) or image_has_competitor_brand),
        "analysis": {
            "source": analysis.source,
            "packageType": analysis.package_type,
            "keywords": analysis.keywords,
            "brands": analysis.brands,
            "productText": analysis.product_text,
            "isMultiProduct": analysis.is_multi_product,
            "primaryProductDescription": analysis.primary_product_description,
            "manualHints": _split_hints(hints),
            "apiKeySource": analysis.api_key_source,
            "modelProfile": analysis.model_profile,
            "modelCalls": analysis.model_calls or [],
            "hasCompetitorBrand": image_has_competitor_brand,
            "footerCropped": footer_cropped,
            "visionRetried": vision_retried,
            "matchedSnowBrands": matched_snow_brands,
            "matchedCompetitorBrands": matched_competitor_brands,
        },
        "pipeline": {
            "stage1": {
                "name": "瓶/听/箱",
                "packageType": pipeline.stage1_package_type,
                "candidateCount": pipeline.stage1_count,
                "fallback": pipeline.stage1_fallback,
            },
            "stage2": {
                "name": "命中品牌强约束 + 关键词",
                "keywords": [*pipeline.stage2_keywords, *(["多品牌混合图，建议人工审核"] if image_has_competitor_brand else [])],
                "candidateCount": pipeline.stage2_count,
                "fallback": pipeline.stage2_fallback,
            },
            "stage3": {
                "name": "图片语义相似度",
                "candidateCount": pipeline.stage3_count,
            },
        },
        "models": {
            "embeddingModel": model_settings["embedding_model"],
            "vlModel": model_settings["vl_model"],
            "enableVlRerank": model_settings["enable_vl_rerank"],
            "qwenTimeoutSeconds": model_settings["qwen_timeout_seconds"],
        },
    }
    passed_gate_reason = f"雪花体系品牌通过：{'、'.join(matched_snow_brands)}，进入模板匹配"
    _write_recognition_log(
        filename=filename,
        trace_id=effective_trace_id,
        result=result,
        analysis_source=analysis.source,
        cleaned_texts=cleaned_analysis_texts,
        gate_reason=passed_gate_reason,
        image_has_snow_brand=image_has_snow_brand,
        image_has_competitor_brand=image_has_competitor_brand,
        matched_snow_brands=matched_snow_brands,
        matched_competitor_brands=matched_competitor_brands,
        model_calls=analysis.model_calls or [],
        footer_cropped=footer_cropped,
    )
    return result


def _analyze_product_image_with_retry(
    upload_path: Path,
    api_key_override: str | None = None,
    model_profile: str = "default",
):
    analysis = analyze_product_image(upload_path, api_key_override=api_key_override, model_profile=model_profile)
    if analysis.source not in {"model_error", "model_timeout", "qwen_error", "qwen_timeout"}:
        return analysis, False
    retry_analysis = analyze_product_image(upload_path, api_key_override=api_key_override, model_profile=model_profile)
    return (
        retry_analysis.__class__(**{**retry_analysis.__dict__, "model_calls": [*(analysis.model_calls or []), *(retry_analysis.model_calls or [])]}),
        True,
    )


def _recognize_with_local_visual_fallback(
    *,
    recognition_path: Path,
    filename: str,
    trace_id: str,
    hints: str,
    analysis_source: str,
    api_key_source: str,
    footer_cropped: bool,
    vision_retried: bool,
    model_calls: list[dict[str, object]],
    model_settings: dict[str, object],
) -> dict[str, object]:
    ranking_weights = {
        key: float(model_settings[key])
        for key in [
            "weight_brand_match_visual",
            "weight_brand_match_brand",
            "weight_brand_match_text",
            "weight_brand_miss_visual",
            "weight_brand_miss_text",
            "weight_no_brand_visual",
            "weight_no_brand_text",
        ]
    }
    hint_keywords = _split_hints(hints)
    matches, pipeline = match_uploaded_image(
        upload_path=recognition_path,
        index_path=settings.vector_index_path,
        template_images_dir=settings.template_images_dir,
        top_k=int(model_settings.get("top_k", 5)),
        package_type=None,
        brands=[],
        keywords=hint_keywords,
        weights=ranking_weights,
        strict_brand_filter=False,
    )
    threshold = float(model_settings["low_confidence_threshold"])
    best = matches[0] if matches else None
    source = f"{analysis_source}_local_fallback"
    result = {
        "filename": filename,
        "traceId": trace_id,
        "best": best.__dict__ if best else None,
        "candidates": [match.__dict__ for match in matches],
        "reviewRequired": True if best else False,
        "analysis": {
            "source": source,
            "packageType": None,
            "keywords": hint_keywords,
            "brands": [],
            "productText": [],
            "isMultiProduct": False,
            "primaryProductDescription": "模型识别失败，已使用本地视觉相似度兜底。",
            "manualHints": hint_keywords,
            "apiKeySource": api_key_source,
            "modelProfile": "local_fallback",
            "modelCalls": model_calls,
            "hasCompetitorBrand": False,
            "footerCropped": footer_cropped,
            "visionRetried": vision_retried,
            "matchedSnowBrands": [],
            "matchedCompetitorBrands": [],
        },
        "pipeline": {
            "stage1": {
                "name": "本地视觉兜底",
                "packageType": pipeline.stage1_package_type,
                "candidateCount": pipeline.stage1_count,
                "fallback": True,
            },
            "stage2": {
                "name": "人工提示关键词",
                "keywords": [*pipeline.stage2_keywords, "模型失败后本地视觉兜底"],
                "candidateCount": pipeline.stage2_count,
                "fallback": pipeline.stage2_fallback,
            },
            "stage3": {
                "name": "图片语义相似度",
                "candidateCount": pipeline.stage3_count,
            },
        },
        "models": {
            "embeddingModel": model_settings["embedding_model"],
            "vlModel": model_settings["vl_model"],
            "enableVlRerank": model_settings["enable_vl_rerank"],
            "qwenTimeoutSeconds": model_settings["qwen_timeout_seconds"],
        },
    }
    result["reviewRequired"] = bool(best and best.score < threshold) or bool(best)
    gate_reason = f"模型识别失败（{analysis_source}），已重试{'一次' if vision_retried else '0次'}并使用本地视觉兜底"
    _write_recognition_log(
        filename=filename,
        trace_id=trace_id,
        result=result,
        analysis_source=source,
        cleaned_texts=[],
        gate_reason=gate_reason,
        image_has_snow_brand=False,
        image_has_competitor_brand=False,
        matched_snow_brands=[],
        matched_competitor_brands=[],
        model_calls=model_calls,
        footer_cropped=footer_cropped,
    )
    return result


def _candidate_reason(candidate: dict[str, object]) -> str:
    matched = candidate.get("matched_keywords") or []
    matched_text = "、".join(str(item) for item in matched[:8]) if isinstance(matched, list) else ""
    parts = [
        f"融合分 {float(candidate.get('score') or 0):.3f}",
        f"视觉 {float(candidate.get('visual_score') or 0):.3f}",
        f"品牌 {float(candidate.get('brand_score') or 0):.3f}",
        f"关键词 {float(candidate.get('text_score') or 0):.3f}",
    ]
    if matched_text:
        parts.append(f"命中词：{matched_text}")
    return "；".join(parts)


def _write_recognition_log(
    *,
    filename: str,
    trace_id: str,
    result: dict[str, object],
    analysis_source: str,
    cleaned_texts: list[str],
    gate_reason: str,
    image_has_snow_brand: bool,
    image_has_competitor_brand: bool,
    matched_snow_brands: list[str],
    matched_competitor_brands: list[str],
    model_calls: list[dict[str, object]],
    footer_cropped: bool,
) -> None:
    best = result.get("best")
    candidates = result.get("candidates") if isinstance(result.get("candidates"), list) else []
    candidate_logs = []
    for candidate in candidates[:5]:
        if not isinstance(candidate, dict):
            continue
        candidate_logs.append(
            {
                "productCode": candidate.get("product_code", ""),
                "productName": candidate.get("product_name", ""),
                "packageType": candidate.get("package_type", ""),
                "view": candidate.get("view", ""),
                "templateImageUrl": candidate.get("template_image_url"),
                "score": candidate.get("score"),
                "visualScore": candidate.get("visual_score"),
                "brandScore": candidate.get("brand_score"),
                "textScore": candidate.get("text_score"),
                "matchedKeywords": candidate.get("matched_keywords", []),
                "reason": _candidate_reason(candidate),
            }
        )
    append_recognition_log(
        {
            "filename": filename,
            "traceId": trace_id,
            "status": "命中" if best else "未命中",
            "analysisSource": analysis_source,
            "gate": {
                "hasSnowBrand": image_has_snow_brand,
                "hasCompetitorBrand": image_has_competitor_brand,
                "matchedSnowBrands": matched_snow_brands,
                "matchedCompetitorBrands": matched_competitor_brands,
                "footerCropped": footer_cropped,
                "reason": gate_reason,
                "cleanedTexts": cleaned_texts[:20],
            },
            "pipeline": result.get("pipeline", {}),
            "modelCalls": model_calls,
            "best": candidate_logs[0] if best and candidate_logs else None,
            "candidates": candidate_logs,
            "reviewRequired": result.get("reviewRequired", False),
            "matchReason": _candidate_reason(best) if isinstance(best, dict) else gate_reason,
        }
    )


async def recognize_upload(
    file: UploadFile,
    hints: str = "",
    api_key_override: str | None = None,
    trace_id: str = "",
    model_profile: str = "default",
) -> dict[str, object]:
    suffix = _validate_upload(file)
    upload_dir = settings.data_dir / "uploads" / "single"
    upload_dir.mkdir(parents=True, exist_ok=True)
    upload_path = upload_dir / f"{uuid.uuid4().hex}{suffix}"

    with upload_path.open("wb") as output:
        shutil.copyfileobj(file.file, output)

    return _recognize_saved_image(
        upload_path,
        file.filename or upload_path.name,
        hints=hints,
        api_key_override=api_key_override,
        trace_id=trace_id,
        model_profile=model_profile,
    )


def _suffix_from_url_or_content_type(image_url: str, content_type: str | None) -> str:
    suffix = Path(urlparse(image_url).path).suffix.lower()
    if suffix in SUPPORTED_EXTENSIONS:
        return suffix
    content_type_map = {
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
    }
    if content_type and content_type.split(";")[0].lower() in content_type_map:
        return content_type_map[content_type.split(";")[0].lower()]
    raise HTTPException(status_code=400, detail="图片链接必须指向 JPG、PNG 或 WebP 图片")


def recognize_image_url(
    image_url: str,
    hints: str = "",
    api_key_override: str | None = None,
    trace_id: str = "",
    model_profile: str = "default",
) -> dict[str, object]:
    parsed = urlparse(image_url)
    if parsed.scheme not in {"http", "https"}:
        raise HTTPException(status_code=400, detail="图片链接仅支持 http 或 https")

    suffix, data = _download_image_url_with_retry(image_url)

    if len(data) > settings.upload_max_mb * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"图片不能超过 {settings.upload_max_mb}MB")

    upload_dir = settings.data_dir / "uploads" / "url"
    upload_dir.mkdir(parents=True, exist_ok=True)
    upload_path = upload_dir / f"{uuid.uuid4().hex}{suffix}"
    upload_path.write_bytes(data)

    return _recognize_saved_image(
        upload_path,
        image_url,
        hints=hints,
        api_key_override=api_key_override,
        trace_id=trace_id,
        model_profile=model_profile,
    )


def _download_image_url_with_retry(image_url: str) -> tuple[str, bytes]:
    last_error: Exception | None = None
    for attempt in range(2):
        request = Request(
            image_url,
            headers={"User-Agent": "ImageRecognition/0.1"},
            method="GET",
        )
        try:
            with urlopen(request, timeout=20) as response:
                content_type = response.headers.get("Content-Type")
                suffix = _suffix_from_url_or_content_type(image_url, content_type)
                max_bytes = settings.upload_max_mb * 1024 * 1024
                return suffix, response.read(max_bytes + 1)
        except HTTPError as exc:
            last_error = exc
            if exc.code not in RETRIABLE_IMAGE_DOWNLOAD_STATUS or attempt >= 1:
                raise HTTPException(status_code=400, detail=f"图片链接下载失败：HTTP {exc.code}") from exc
        except URLError as exc:
            last_error = exc
            if attempt >= 1:
                raise HTTPException(status_code=400, detail=f"图片链接下载失败：{exc.reason}") from exc
        except TimeoutError as exc:
            last_error = exc
            if attempt >= 1:
                raise HTTPException(status_code=400, detail="图片链接下载超时") from exc
        time.sleep(1)
    raise HTTPException(status_code=400, detail=f"图片链接下载失败：{last_error}")


def _normalize_trace_id(value: object) -> str:
    text = str(value or "").strip()
    return "".join(char for char in text if char.isalnum() or char in {"-", "_", "."})[:80]


def _generate_trace_id() -> str:
    return f"R{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
