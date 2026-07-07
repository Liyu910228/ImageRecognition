from io import BytesIO

from fastapi import APIRouter, File, Form, Header, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, HttpUrl

from app.features.recognition.service import extract_image_links_from_workbook, recognize_image_url, recognize_upload


class RecognizeUrlRequest(BaseModel):
    image_url: HttpUrl
    hints: str = ""
    trace_id: str = ""
    model_profile: str = "default"


class BatchExportRow(BaseModel):
    trace_id: str = ""
    source: str
    status: str
    product_code: str = ""
    product_name: str = ""
    package_type: str = ""
    score: float | None = None
    sheet: str = ""
    row: int | None = None
    error: str = ""


class BatchExportRequest(BaseModel):
    rows: list[BatchExportRow]


router = APIRouter()


@router.post("/recognize")
async def recognize(
    file: UploadFile = File(...),
    hints: str = Form(default=""),
    trace_id: str = Form(default=""),
    model_profile: str = Form(default="default"),
    dashscope_api_key: str | None = Header(default=None, alias="X-DashScope-Api-Key"),
) -> dict[str, object]:
    return await recognize_upload(
        file,
        hints=hints,
        api_key_override=dashscope_api_key,
        trace_id=trace_id,
        model_profile=model_profile,
    )


@router.post("/recognize-url")
def recognize_url(
    payload: RecognizeUrlRequest,
    dashscope_api_key: str | None = Header(default=None, alias="X-DashScope-Api-Key"),
) -> dict[str, object]:
    return recognize_image_url(
        str(payload.image_url),
        hints=payload.hints,
        api_key_override=dashscope_api_key,
        trace_id=payload.trace_id,
        model_profile=payload.model_profile,
    )


@router.post("/recognize-url-workbook")
def recognize_url_workbook(file: UploadFile = File(...)) -> dict[str, object]:
    return extract_image_links_from_workbook(file)


@router.post("/batch-results/export")
def export_batch_results(payload: BatchExportRequest) -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "批量识别结果"
    headers = ["识别编号", "来源", "状态", "产品编码", "产品名称", "瓶/听/箱", "相似度"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="DDEFEA")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="143B36")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in payload.rows:
        sheet.append(
            [
                item.trace_id,
                item.source,
                item.status,
                item.product_code,
                item.product_name,
                item.package_type,
                None if item.score is None else round(item.score * 100, 1),
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        row[6].number_format = '0.0"%"'
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_widths = [26, 56, 12, 18, 52, 12, 12]
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="batch-recognition-results.xlsx"'},
    )
