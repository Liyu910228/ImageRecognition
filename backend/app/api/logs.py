from io import BytesIO

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.features.recognition.logs import clear_recognition_logs, read_recognition_logs


router = APIRouter()


@router.get("/logs/recognition")
def list_recognition_logs(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200, alias="pageSize"),
    q: str = Query(default=""),
    status: str = Query(default="all"),
    snow_gate: str = Query(default="all", alias="snowGate"),
    competitor: str = Query(default="all"),
) -> dict[str, object]:
    page_data = read_recognition_logs(
        limit=page_size,
        page=page,
        query=q,
        status=status,
        snow_gate=snow_gate,
        competitor=competitor,
    )
    return {"logs": page_data["items"], "total": page_data["total"], "page": page_data["page"], "pageSize": page_data["pageSize"]}


@router.delete("/logs/recognition")
def delete_recognition_logs() -> dict[str, object]:
    clear_recognition_logs()
    return {"ok": True}


@router.get("/logs/recognition/export")
def export_recognition_logs(
    q: str = Query(default=""),
    status: str = Query(default="all"),
    snow_gate: str = Query(default="all", alias="snowGate"),
    competitor: str = Query(default="all"),
    ids: str = Query(default=""),
) -> StreamingResponse:
    selected_ids = {item.strip() for item in ids.split(",") if item.strip()} or None
    page_data = read_recognition_logs(
        limit=100000,
        page=1,
        query=q,
        status=status,
        snow_gate=snow_gate,
        competitor=competitor,
        selected_ids=selected_ids,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "recognition-logs"
    headers = [
        "日志序号",
        "识别编号",
        "识别时间",
        "文件",
        "状态",
        "识别来源",
        "门禁原因",
        "最终原因",
        "雪花品牌门禁",
        "命中雪花品牌",
        "竞品/混合图",
        "命中竞品品牌",
        "底部文字裁剪",
        "候选排名",
        "候选产品编码",
        "候选产品名称",
        "瓶/听/箱",
        "视角",
        "相似度",
        "视觉分",
        "品牌分",
        "关键词分",
        "命中关键词",
        "候选模板图",
        "候选原因",
        "模型调用",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="DDEFEA")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="143B36")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for log_index, log in enumerate(page_data["items"], start=1):
        candidates = log.get("candidates") if isinstance(log.get("candidates"), list) else []
        if not candidates:
            candidates = [None]
        gate = log.get("gate") if isinstance(log.get("gate"), dict) else {}
        for candidate_index, candidate in enumerate(candidates, start=1):
            candidate_data = candidate if isinstance(candidate, dict) else {}
            sheet.append(
                [
                    log_index,
                    log.get("traceId", ""),
                    log.get("createdAt", ""),
                    log.get("filename", ""),
                    _status_label(str(log.get("status", ""))),
                    log.get("analysisSource", ""),
                    gate.get("reason", ""),
                    log.get("matchReason", ""),
                    "通过" if gate.get("hasSnowBrand") else "未通过",
                    " / ".join(str(item) for item in gate.get("matchedSnowBrands", []) or []),
                    "是" if gate.get("hasCompetitorBrand") else "否",
                    " / ".join(str(item) for item in gate.get("matchedCompetitorBrands", []) or []),
                    "是" if gate.get("footerCropped") else "否",
                    candidate_index if candidate else "",
                    candidate_data.get("productCode", ""),
                    candidate_data.get("productName", ""),
                    candidate_data.get("packageType", ""),
                    candidate_data.get("view", ""),
                    _percent(candidate_data.get("score")),
                    _percent(candidate_data.get("visualScore")),
                    _percent(candidate_data.get("brandScore")),
                    _percent(candidate_data.get("textScore")),
                    " / ".join(str(item) for item in candidate_data.get("matchedKeywords", []) or []),
                    candidate_data.get("templateImageUrl", ""),
                    candidate_data.get("reason", ""),
                    _format_model_calls(log.get("modelCalls")),
                ]
            )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [10, 24, 24, 48, 12, 20, 42, 42, 14, 24, 14, 24, 14, 10, 18, 46, 12, 12, 12, 12, 12, 12, 30, 48, 54, 54]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="recognition-logs.xlsx"'},
    )


def _status_label(value: str) -> str:
    if value in {"命中", "鍛戒腑"}:
        return "命中"
    if value in {"未命中", "鏈懡涓?"} or value.startswith("鏈"):
        return "未命中"
    return value


def _percent(value: object) -> float | None:
    try:
        return round(float(value) * 100, 2)
    except (TypeError, ValueError):
        return None


def _format_model_calls(value: object) -> str:
    if not isinstance(value, list):
        return ""
    parts = []
    for item in value:
        if not isinstance(item, dict):
            continue
        parts.append(f"{item.get('role', '')}:{item.get('model', '')} {item.get('source', '')} {item.get('elapsedMs', '')}ms")
    return " / ".join(parts)
